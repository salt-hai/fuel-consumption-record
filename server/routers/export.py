from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from typing import Optional
from datetime import datetime
import csv
import io
from database import get_db
from models.fuel_record import FuelRecord
from models.vehicle import Vehicle
from models.user import User
from utils.auth import get_current_user

router = APIRouter(prefix="/v1/export", tags=["数据导出"])

async def get_export_data(
    vehicle_id: Optional[int],
    start_date: Optional[str],
    end_date: Optional[str],
    user: User,
    db: AsyncSession
):
    """获取当前用户的导出数据"""
    # 只查询当前用户车辆的记录
    query = select(FuelRecord).join(Vehicle, FuelRecord.vehicle_id == Vehicle.id).where(Vehicle.user_id == user.id)

    if vehicle_id:
        # 验证车辆属于当前用户
        result = await db.execute(
            select(Vehicle).where(Vehicle.id == vehicle_id, Vehicle.user_id == user.id)
        )
        if not result.scalar_one_or_none():
            return []  # 车辆不存在或不属于当前用户
        query = query.where(FuelRecord.vehicle_id == vehicle_id)
    if start_date:
        query = query.where(FuelRecord.date >= start_date)
    if end_date:
        query = query.where(FuelRecord.date <= end_date)

    query = query.order_by(FuelRecord.date.desc())
    query = query.options(selectinload(FuelRecord.vehicle))
    result = await db.execute(query)
    return result.scalars().all()

@router.get("/csv")
async def export_csv(
    vehicle_id: Optional[int] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """导出当前用户的加油记录为 CSV"""
    records = await get_export_data(vehicle_id, start_date, end_date, current_user, db)

    output = io.StringIO()
    writer = csv.writer(output)

    # 写入表头
    writer.writerow(["车牌号", "车辆名称", "日期", "里程(km)", "加油量(L)", "总金额(元)", "单价(元/L)", "是否加满", "加油站", "油耗(L/100km)", "备注"])

    # 写入数据
    for r in records:
        writer.writerow([
            r.vehicle.plate_number if r.vehicle else "",
            r.vehicle.name if r.vehicle else "",
            r.date,
            r.odometer,
            r.volume,
            r.total_cost,
            r.unit_price or "",
            "是" if r.full_tank else "否",
            r.gas_station or "",
            f"{r.fuel_consumption:.1f}" if r.fuel_consumption else "",
            r.notes or ""
        ])

    output.seek(0)
    filename = f"fuel_records_{datetime.now().strftime('%Y%m%d%H%M%S')}.csv"

    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),  # UTF-8 BOM for Excel
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/excel")
async def export_excel(
    vehicle_id: Optional[int] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """导出当前用户的加油记录为 Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    records = await get_export_data(vehicle_id, start_date, end_date, current_user, db)

    wb = Workbook()
    ws = wb.active
    ws.title = "加油记录"

    # 表头
    headers = ["车牌号", "车辆名称", "日期", "里程(km)", "加油量(L)", "总金额(元)", "单价(元/L)", "是否加满", "加油站", "油耗(L/100km)", "备注"]
    ws.append(headers)

    # 表头样式
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 数据行
    for r in records:
        ws.append([
            r.vehicle.plate_number if r.vehicle else "",
            r.vehicle.name if r.vehicle else "",
            r.date,
            r.odometer,
            r.volume,
            r.total_cost,
            r.unit_price or "",
            "是" if r.full_tank else "否",
            r.gas_station or "",
            f"{r.fuel_consumption:.1f}" if r.fuel_consumption else "",
            r.notes or ""
        ])

    # 调整列宽
    ws.column_dimensions['A'].width = 12  # 车牌号
    ws.column_dimensions['B'].width = 12  # 车辆名称
    ws.column_dimensions['C'].width = 12  # 日期
    ws.column_dimensions['D'].width = 12  # 里程
    ws.column_dimensions['E'].width = 10  # 加油量
    ws.column_dimensions['F'].width = 10  # 总金额
    ws.column_dimensions['G'].width = 10  # 单价
    ws.column_dimensions['H'].width = 10  # 是否加满
    ws.column_dimensions['I'].width = 15  # 加油站
    ws.column_dimensions['J'].width = 12  # 油耗
    ws.column_dimensions['K'].width = 20  # 备注

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"fuel_records_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
